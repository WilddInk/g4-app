import argparse
import json
import os
from pathlib import Path
from urllib import parse, request

import openpyxl


def load_env_file(path: Path) -> dict:
    env = {}
    if not path.exists():
        return env
    for line in path.read_text(encoding="utf-8").splitlines():
        raw = line.strip()
        if not raw or raw.startswith("#") or "=" not in raw:
            continue
        k, v = raw.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def text(v):
    if v is None:
        return ""
    return str(v).strip()


def pick_first_nonempty(*vals):
    for v in vals:
        t = text(v)
        if t:
            return t
    return ""


def to_bool(v):
    if isinstance(v, bool):
        return v
    t = text(v).lower()
    if t in {"1", "true", "tak", "yes", "y"}:
        return True
    if t in {"0", "false", "nie", "no", "n"}:
        return False
    return None


def parse_xlsx(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [text(ws.cell(row=1, column=i).value) for i in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    out = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=c).value for h, c in idx.items()}
        nr = text(row.get("ID"))
        if not nr or nr.lower() in {"id", "nr"}:
            continue

        imie = text(row.get("Imię")) or text(row.get("Imi�"))
        nazwisko = text(row.get("Nazwisko"))
        imie_nazwisko = " ".join([x for x in [imie, nazwisko] if x]).strip()

        telefon = pick_first_nonempty(row.get("Telefon Firmowy"), row.get("Telefon Prywatny"))
        email = pick_first_nonempty(row.get("Mail Firmowy"), row.get("Mail Gmail"), row.get("Mail Prywatny"))

        status = text(row.get("Status")).lower()
        active_from_status = None
        if status:
            if "aktywn" in status:
                active_from_status = True
            elif any(x in status for x in ["nieaktywn", "archiw", "zwoln", "usun"]):
                active_from_status = False
        is_active = True if active_from_status is None else bool(active_from_status)

        # Skip known template/header row copied into data.
        if nr == "000" and imie.lower().startswith("imi") and nazwisko.lower().startswith("nazw"):
            continue

        record = {
            "nr": nr,
            "imie_nazwisko": imie_nazwisko if imie_nazwisko else nr,
            "telefon": telefon or None,
            "email": email or None,
            "is_active": bool(is_active),
        }
        out.append(record)
    return out


def upsert_pracownicy(url: str, service_key: str, rows):
    endpoint = f"{url.rstrip('/')}/rest/v1/pracownik?on_conflict=nr"
    body = json.dumps(rows).encode("utf-8")
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    req = request.Request(endpoint, data=body, headers=headers, method="POST")
    with request.urlopen(req) as resp:
        return resp.status


def sql_literal(v):
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "true" if v else "false"
    s = str(v).replace("'", "''")
    return f"'{s}'"


def write_sql_file(rows, out_path: Path):
    lines = []
    lines.append("-- Auto-generated from XLSX: import_pracownik_from_xlsx.py")
    lines.append("-- Columns limited to existing app fields: nr, imie_nazwisko, telefon, email, is_active")
    lines.append("BEGIN;")
    lines.append("")
    lines.append("CREATE TEMP TABLE tmp_pracownik_import (")
    lines.append("  nr text,")
    lines.append("  imie_nazwisko text,")
    lines.append("  telefon text,")
    lines.append("  email text,")
    lines.append("  is_active boolean")
    lines.append(") ON COMMIT DROP;")
    lines.append("")
    lines.append("INSERT INTO tmp_pracownik_import (nr, imie_nazwisko, telefon, email, is_active)")
    lines.append("VALUES")
    vals = []
    for r in rows:
        vals.append(
            f"  ({sql_literal(r['nr'])}, {sql_literal(r['imie_nazwisko'])}, {sql_literal(r['telefon'])}, {sql_literal(r['email'])}, {sql_literal(r['is_active'])})"
        )
    lines.append(",\n".join(vals))
    lines.append(";")
    lines.append("")
    lines.append("INSERT INTO public.pracownik (nr, imie_nazwisko, telefon, email, is_active)")
    lines.append("SELECT")
    lines.append("  s.nr,")
    lines.append("  s.imie_nazwisko,")
    lines.append("  s.telefon,")
    lines.append("  CASE")
    lines.append("    WHEN NULLIF(trim(coalesce(s.email, '')), '') IS NULL THEN NULL")
    lines.append("    WHEN EXISTS (")
    lines.append("      SELECT 1")
    lines.append("      FROM tmp_pracownik_import d")
    lines.append("      WHERE d.nr <> s.nr")
    lines.append("        AND lower(trim(coalesce(d.email, ''))) = lower(trim(coalesce(s.email, '')))")
    lines.append("    ) THEN NULL")
    lines.append("    WHEN EXISTS (")
    lines.append("      SELECT 1")
    lines.append("      FROM public.pracownik p")
    lines.append("      WHERE p.nr <> s.nr")
    lines.append("        AND lower(trim(coalesce(p.email, ''))) = lower(trim(coalesce(s.email, '')))")
    lines.append("    ) THEN NULL")
    lines.append("    ELSE s.email")
    lines.append("  END AS email,")
    lines.append("  s.is_active")
    lines.append("FROM tmp_pracownik_import s")
    lines.append("ON CONFLICT (nr) DO UPDATE SET")
    lines.append("  imie_nazwisko = EXCLUDED.imie_nazwisko,")
    lines.append("  telefon = EXCLUDED.telefon,")
    lines.append("  email = EXCLUDED.email,")
    lines.append("  is_active = EXCLUDED.is_active;")
    lines.append("")
    lines.append("COMMIT;")
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    ap = argparse.ArgumentParser(description="Import pracownik from xlsx into Supabase")
    ap.add_argument("--xlsx", required=True, help="Path to xlsx file")
    ap.add_argument("--apply", action="store_true", help="Actually write to DB")
    ap.add_argument("--sql-output", help="Write UPSERT SQL file instead of API write")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    rows = parse_xlsx(xlsx_path)
    active = sum(1 for r in rows if r.get("is_active") is True)
    inactive = len(rows) - active
    print(f"Parsed rows: {len(rows)} (active={active}, inactive={inactive})")
    print("Sample:", json.dumps(rows[:5], ensure_ascii=False))

    if args.sql_output:
        out_path = Path(args.sql_output)
        write_sql_file(rows, out_path)
        print(f"SQL file written: {out_path}")
        return

    if not args.apply:
        print("Dry-run only. Use --apply to write to Supabase, or --sql-output for SQL file.")
        return

    env_local = load_env_file(Path(__file__).resolve().parents[1] / ".env")
    supabase_url = (
        os.getenv("SUPABASE_URL")
        or env_local.get("SUPABASE_URL", "")
        or os.getenv("VITE_SUPABASE_URL")
        or env_local.get("VITE_SUPABASE_URL", "")
    )
    service_key = (
        os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        or env_local.get("SUPABASE_SERVICE_ROLE_KEY", "")
        or os.getenv("VITE_SUPABASE_ANON_KEY")
        or env_local.get("VITE_SUPABASE_ANON_KEY", "")
    )
    if not supabase_url or not service_key:
        raise SystemExit("Missing Supabase URL/key in env/.env")

    status = upsert_pracownicy(supabase_url, service_key, rows)
    print(f"Upsert completed. HTTP {status}")


if __name__ == "__main__":
    main()

