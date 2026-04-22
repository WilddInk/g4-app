#!/usr/bin/env python3
"""
Generuje plik SQL (INSERT do public.czas_pracy_wpis) z arkusza miesięcznego
„RozliczenieGodzinowe” z CMP (arkusze MM-RRRR, bez „RozliczenieKosztów”).

Użycie:
  python import_rozliczenie_godzinowe_xlsx_to_sql.py sciezka/do/pliku.xlsx [--out supabase/import-....sql]

Wymaga: openpyxl
"""

from __future__ import annotations

import argparse
import calendar
import pathlib
import re
from datetime import date, datetime, time

import openpyxl

MM_RRRR = re.compile(r"^\d{2}-\d{4}$")

# KR ogólnobiurowa (biuro) — w Excelu często jako 0 liczbowe lub „0”.
KR_OGOLNOBIUROWY = "000"


def _dopelnij_kr_trzy_cyfry_leading_zero(s: str) -> str:
    """Trzy cyfry → dopisz 0 z przodu (779→0779); „000” bez zmian."""
    if not s or s == KR_OGOLNOBIUROWY:
        return s
    return "0" + s if re.fullmatch(r"\d{3}", s) else s


def normalizuj_kr_z_arkusza(kr_val) -> str:
    if kr_val is None or kr_val == "":
        return ""
    if isinstance(kr_val, (int, float)) and not isinstance(kr_val, bool) and float(kr_val) == 0:
        return KR_OGOLNOBIUROWY
    s = str(kr_val).strip()
    if s == "0":
        return KR_OGOLNOBIUROWY
    return _dopelnij_kr_trzy_cyfry_leading_zero(s)


def _unikalne_zachowujac_kolejnosc(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for x in items:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def rozbij_kry_z_komorki(kr_val) -> list[str]:
    """Kilka KR w jednej komórce E (np. „000/1079”) → osobne wpisy przed podziałem godzin."""
    if kr_val is None or kr_val == "":
        return []
    if isinstance(kr_val, (int, float)) and not isinstance(kr_val, bool) and float(kr_val) == 0:
        return [KR_OGOLNOBIUROWY]
    raw = str(kr_val).strip()
    if not raw:
        return []
    if raw == "0":
        return [KR_OGOLNOBIUROWY]

    po_sep = [normalizuj_kr_z_arkusza(p.strip()) for p in re.split(r"[,;/|]+", raw)]
    po_sep = [p for p in po_sep if p]
    if len(po_sep) > 1:
        return _unikalne_zachowujac_kolejnosc(po_sep)

    multi = re.findall(r"\d{3,4}", raw)
    wyglada_na_wiele = multi and len(multi) > 1 and (bool(re.search(r"[^\d\s]", raw)) or bool(re.search(r"\d\s+\d", raw)))
    if wyglada_na_wiele:
        nums = [normalizuj_kr_z_arkusza(m) for m in multi]
        nums = [n for n in nums if n]
        return _unikalne_zachowujac_kolejnosc(nums)

    one = normalizuj_kr_z_arkusza(kr_val)
    return [one] if one else []


def sql_str(s: str | None) -> str:
    if s is None:
        return "NULL"
    t = str(s).replace("'", "''")
    return "'" + t + "'"


def make_sql_dollar():
    """Factory: unikatowy tag $…$ na każde pole (treść nie może zawierać `$tag$`)."""
    n = 0

    def sql_dollar(s: str | None) -> str:
        nonlocal n
        if s is None:
            return "NULL"
        body = str(s)
        for _ in range(500):
            n += 1
            tag = f"cg{n}"
            delim = f"${tag}$"
            if delim not in body:
                return delim + body + delim
        raise ValueError("Nie znaleziono delimiterów dollar-quote dla tekstu")

    return sql_dollar


def sql_num(n: float | None) -> str:
    if n is None:
        return "NULL"
    return f"{float(n):.2f}"


def format_time_cell(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, datetime):
        return v.strftime("%H:%M")
    return str(v).strip() or None


def _wartosc_z_komorki_ewidencyjnej(v) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return str(int(v))
    s = str(v).strip()
    if s.isdigit():
        return s
    m = re.search(r"\d+", s)
    return m.group(0) if m else None


def _priorytet_etykiety_numeru_ewid(label_raw: str) -> int | None:
    t = str(label_raw).strip()
    low = t.lower()
    if "numer ewid" not in low and "numer ewident" not in low:
        return None
    if re.match(r"^numer\s+ewid\.?\s*$", t, re.I):
        return 320
    if re.match(r"^numer ewidencyjny\s*:?\s*$", t, re.I):
        return 310
    if re.match(r"^numer ewidencyjny\b", t, re.I) and len(t) <= 42:
        return 280
    if re.match(r"^numer\s+ewid\b", t, re.I) and len(t) <= 36:
        return 260
    return max(0, 140 - min(len(t), 140))


def znajdz_numer_ewidencyjny(ws, max_scan: int = 30) -> str | None:
    """Wartość z kolumny D lub (gdy pusto) E — priorytet dla standardowej krótkiej etykiety CMP."""
    cand: list[tuple[int, str, int]] = []
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        b = row[1] if len(row) > 1 else None
        if b is None:
            continue
        prio = _priorytet_etykiety_numeru_ewid(str(b))
        if prio is None:
            continue
        d = row[3] if len(row) > 3 else None
        e = row[4] if len(row) > 4 else None
        nr = _wartosc_z_komorki_ewidencyjnej(d)
        if not nr:
            nr = _wartosc_z_komorki_ewidencyjnej(e)
        if not nr:
            continue
        cand.append((idx, nr, prio))
    if not cand:
        return None
    cand.sort(key=lambda x: (-x[2], x[0]))
    return cand[0][1]


def znajdz_wiersz_naglowka(ws, max_scan: int = 40) -> int | None:
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        b = row[1] if len(row) > 1 else None
        if b is None:
            continue
        if str(b).strip().upper() == "DATA":
            return idx
    return None


def czy_pomin_j(raw_j) -> bool:
    if raw_j is None:
        return True
    if isinstance(raw_j, str):
        t = raw_j.strip()
        if not t:
            return True
        if "wpisz czas" in t.lower():
            return True
        return False
    return False


# Domyślna długość „pełnego dnia” przy nieobecnościach z arkusza (brak osobnej liczby godzin).
GODZ_PELNY_DZIEN_ARKUSZ = 8.0


def mapuj_kod_arkusza(j_raw: str) -> tuple[str, float, list[str]]:
    """
    Mapuje tekst z kolumny J (Excel CMP) na wartość `typ` z aplikacji (CzasPracyPanel / CZAS_TYP_WPISU).
    Zwraca (typ, godziny, dodatkowe_linie_do_uwag).
    """
    s0 = (j_raw or "").strip()
    s = s0.replace("–", "-").replace("—", "-")
    sc = re.sub(r"\s+", "", s.lower())

    if re.match(r"^nb-ch-c", sc):
        return (
            "zwolnienie_lekarskie",
            GODZ_PELNY_DZIEN_ARKUSZ,
            ["Kod arkusza: L4 — ciąża (nb-CH-C)"],
        )
    if re.match(r"^nb-ch-o", sc):
        return (
            "zwolnienie_lekarskie",
            GODZ_PELNY_DZIEN_ARKUSZ,
            ["Kod arkusza: L4 — opieka nad dzieckiem, małżonkiem itd. (nb-CH-O)"],
        )
    if re.match(r"^nb-ch", sc):
        return ("zwolnienie_lekarskie", GODZ_PELNY_DZIEN_ARKUSZ, [])
    if re.match(r"^nb-od", sc):
        return ("opieka_nad_dzieckiem", GODZ_PELNY_DZIEN_ARKUSZ, [])
    if re.match(r"^nb-uw", sc):
        return ("urlop", GODZ_PELNY_DZIEN_ARKUSZ, [])
    if re.match(r"^nb-uż", sc) or re.match(r"^nb-uz", sc):
        return ("urlop_na_zadanie", GODZ_PELNY_DZIEN_ARKUSZ, [])
    if re.match(r"^nb-uo", sc):
        return (
            "inne",
            GODZ_PELNY_DZIEN_ARKUSZ,
            ["Kod arkusza: urlop okolicznościowy (nb-UO)"],
        )
    if re.match(r"^nb-wś", sc) or re.match(r"^nb-ws", sc):
        return (
            "inne",
            0.0,
            ["Kod arkusza: wolne za święto w sobotę (nb-WŚ)"],
        )
    if re.match(r"^nb-n", sc):
        return ("inne", 0.0, ["Kod arkusza: niedostępność (nb-N)"])

    if "święto" in sc or "swieto" in sc or sc.startswith("nd-"):
        return ("inne", 0.0, [f"Kod arkusza: {s0}"])

    if sc.startswith("nb-"):
        return ("inne", 0.0, [f"Kod arkusza (niezmapowany): {s0}"])

    return ("inne", 0.0, [f"Kod arkusza: {s0}"])


def ma_zakres_godzin(godz_od, godz_do) -> bool:
    """Oba pola czasu muszą być ustawione — wtedy arkusz rozróżnia zakres dla KR."""
    go = format_time_cell(godz_od)
    gd = format_time_cell(godz_do)
    return bool(go and gd)


def podziel_godziny_rowno(calkowite: float, n: int) -> list[float]:
    """Równy podział sumy godzin na n części (zaokrąglenie setnych)."""
    if n <= 0:
        return []
    setne = round(float(calkowite) * 100)
    baz = setne // n
    rem = setne % n
    out: list[float] = []
    for i in range(n):
        out.append((baz + (1 if i < rem else 0)) / 100)
    return out


UWAGA_PODZIAL_PREFIX = "Import: podział bez zakresu Od–Do —"


def zastosuj_podzial_godzin_miedzy_kr_bez_zakresu(wiersze: list[dict]) -> None:
    """Ten sam dzień, kilka wierszy „praca” bez Od–Do, ta sama wartość godzin, różne KR → dziel równo.
    Wiersze już rozłożone z jednej komórki z wieloma KR są pomijane."""
    from collections import defaultdict

    grupy: dict[str, list[int]] = defaultdict(list)
    for i, w in enumerate(wiersze):
        if (
            w.get("typ") != "praca"
            or not w.get("_raw_j_num")
            or w.get("_zakres")
            or w.get("_juz_podzielone_wielo_kr")
        ):
            continue
        key = f'{w["nr"]}|{w["d_iso"]}'
        grupy[key].append(i)

    for indices in grupy.values():
        if len(indices) < 2:
            continue
        slice_rows = [wiersze[j] for j in indices]
        krs = [str(r.get("kr") or "").strip() for r in slice_rows]
        if any(not k for k in krs):
            continue
        if len(set(krs)) != len(krs):
            continue

        vals = [float(r["gh"]) for r in slice_rows]
        positives = [v for v in vals if v > 0]
        if not positives:
            continue
        uniq_cents = {round(v * 100) for v in positives}
        if len(uniq_cents) != 1:
            continue

        H = positives[0]
        n = len(indices)
        czesci = podziel_godziny_rowno(H, n)

        for j, idx in enumerate(indices):
            wiersze[idx]["gh"] = czesci[j]
            dop = f"{UWAGA_PODZIAL_PREFIX} część {j + 1}/{n}, {H:.2f} h łącznie na dzień przy braku rozróżnienia godzin na KR."
            u = wiersze[idx].get("uwagi") or ""
            wiersze[idx]["uwagi"] = u + "\n" + dop if u else dop

    for w in wiersze:
        w.pop("_zakres", None)
        w.pop("_raw_j_num", None)
        w.pop("_juz_podzielone_wielo_kr", None)


def zakres_miesiaca_z_nazwy_arkusza(sheet_name: str) -> tuple[str, str] | None:
    """Pierwszy i ostatni dzień miesiąca MM-RRRR (ISO) albo None."""
    s = str(sheet_name).strip()
    if not MM_RRRR.match(s):
        return None
    mm_s, yyyy_s = s.split("-", 1)
    y, m = int(yyyy_s), int(mm_s)
    if m < 1 or m > 12:
        return None
    d0 = date(y, m, 1)
    d1 = date(y, m, calendar.monthrange(y, m)[1])
    return d0.isoformat(), d1.isoformat()


def mapuj_wiersz(raw_j, kr: str, nazwa_obiektu, opis, godz_od, godz_do, uwagi_k: str | None) -> tuple[str, float, float, str | None]:
    """Zwraca (typ, godziny, nadgodziny, uwagi_dodatkowe_składowe)."""
    uw_extra_parts: list[str] = []
    if nazwa_obiektu:
        uw_extra_parts.append(f"Obiekt: {str(nazwa_obiektu).strip()}")
    if opis:
        uw_extra_parts.append(str(opis).strip())
    go = format_time_cell(godz_od)
    gd = format_time_cell(godz_do)
    if go or gd:
        uw_extra_parts.append(f"{go or '?'}–{gd or '?'}")
    if uwagi_k:
        uw_extra_parts.append(str(uwagi_k).strip())

    if isinstance(raw_j, (int, float)) and not isinstance(raw_j, bool):
        g = float(raw_j)
        return ("praca", g, 0.0, "\n".join(uw_extra_parts) if uw_extra_parts else None)

    raw_kod = str(raw_j).strip()
    typ, gh, kod_linie = mapuj_kod_arkusza(raw_kod)
    nad = 0.0
    polacz = [*uw_extra_parts, *kod_linie]
    pierwszy = raw_kod.split()[0] if raw_kod else ""
    if pierwszy.lower().startswith("nb-"):
        polacz.append(f"({pierwszy})")
    return (typ, gh, nad, "\n".join(polacz) if polacz else None)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=pathlib.Path, help="Plik .xlsx RozliczenieGodzinowe")
    ap.add_argument(
        "--out",
        type=pathlib.Path,
        default=None,
        help="Plik wyjściowy .sql (domyślnie obok xlsx: nazwa_import.sql)",
    )
    ap.add_argument(
        "--pracownik-nr",
        type=str,
        default=None,
        help="Wymuś numer ewidencyjny zamiast odczytu z arkusza",
    )
    ap.add_argument(
        "--skip-delete",
        action="store_true",
        help="Nie dodawaj DELETE przed INSERT (domyślnie: stare wpisy z tego importu są usuwane).",
    )
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)

    sheets_mm = [name for name in wb.sheetnames if MM_RRRR.match(name.strip())]

    wiersze: list[dict] = []
    meta_nr = args.pracownik_nr
    sql_dollar = make_sql_dollar()
    arkusze_uzyte: set[str] = set()

    for sheet_name in sorted(sheets_mm, key=lambda x: (int(x.split("-")[1]), int(x.split("-")[0]))):
        ws = wb[sheet_name]
        hdr = znajdz_wiersz_naglowka(ws)
        if hdr is None:
            continue
        nr = meta_nr or znajdz_numer_ewidencyjny(ws, hdr - 1)
        if not nr:
            raise SystemExit(f"Brak numeru ewidencyjnego w arkuszu {sheet_name} (szukano przed nagłówkiem).")

        for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
            if len(row) < 11:
                continue
            data_cell = row[1]
            raw_j = row[9]
            if czy_pomin_j(raw_j):
                continue

            if isinstance(data_cell, datetime):
                d_iso = data_cell.date().isoformat()
            elif isinstance(data_cell, date):
                d_iso = data_cell.isoformat()
            else:
                continue

            arkusze_uzyte.add(sheet_name)

            kr_val = row[4]
            kr_lista = rozbij_kry_z_komorki(kr_val)
            krs_do_wierszy = kr_lista if kr_lista else [normalizuj_kr_z_arkusza(kr_val)]

            ma_zakres = ma_zakres_godzin(row[7], row[8])
            raw_j_num = isinstance(raw_j, (int, float)) and not isinstance(raw_j, bool)

            opis_zad = row[6]
            wyk_zad = str(opis_zad).strip() if opis_zad not in (None, "") else None

            n_kr_cell = len(krs_do_wierszy)
            mapped_wspolne_multi_kr = None
            czesci_gh: list[float] | None = None
            if n_kr_cell > 1:
                mapped_wspolne_multi_kr = mapuj_wiersz(
                    raw_j,
                    krs_do_wierszy[0],
                    row[5],
                    row[6],
                    row[7],
                    row[8],
                    row[10] if len(row) > 10 else None,
                )
                if mapped_wspolne_multi_kr[0] == "praca" and isinstance(raw_j, (int, float)) and not isinstance(
                    raw_j, bool
                ):
                    czesci_gh = podziel_godziny_rowno(mapped_wspolne_multi_kr[1], n_kr_cell)

            for j, kr in enumerate(krs_do_wierszy):
                if czesci_gh is not None:
                    typ, _, nad, uw_z_excel = mapped_wspolne_multi_kr
                    mapped_typ = typ
                    gh = czesci_gh[j]
                    juz_wielo = True
                else:
                    typ, gh, nad, uw_z_excel = mapuj_wiersz(
                        raw_j,
                        kr,
                        row[5],
                        row[6],
                        row[7],
                        row[8],
                        row[10] if len(row) > 10 else None,
                    )
                    mapped_typ = typ
                    juz_wielo = False

                if not kr and mapped_typ == "praca":
                    kr = KR_OGOLNOBIUROWY

                suma_h = mapped_wspolne_multi_kr[1] if czesci_gh is not None and mapped_typ == "praca" else gh

                uw_parts: list[str] = []
                if uw_z_excel:
                    uw_parts.append(uw_z_excel)
                uw_parts.append(f"Import: RozliczenieGodzinowe / {sheet_name}")
                if czesci_gh is not None and mapped_typ == "praca":
                    uw_parts.append(
                        f"Import: wiele KR w jednej komórce — część {j + 1}/{n_kr_cell} sumy {suma_h:.2f} h z kolumny „Czas pracy”."
                    )
                uwagi = "\n".join(uw_parts)

                wiersze.append(
                    {
                        "nr": nr,
                        "d_iso": d_iso,
                        "kr": kr,
                        "typ": mapped_typ,
                        "gh": gh,
                        "nad": nad,
                        "uwagi": uwagi,
                        "wyk_zad": wyk_zad,
                        "_zakres": ma_zakres,
                        "_raw_j_num": bool(raw_j_num and mapped_typ == "praca"),
                        "_juz_podzielone_wielo_kr": juz_wielo,
                    }
                )

    wb.close()

    zastosuj_podzial_godzin_miedzy_kr_bez_zakresu(wiersze)

    inserts: list[str] = []
    for w in wiersze:
        inserts.append(
            "INSERT INTO public.czas_pracy_wpis (pracownik_nr, data, kr, typ, godziny, nadgodziny, uwagi, wykonywane_zadanie) "
            f"VALUES ({sql_str(w['nr'])}, '{w['d_iso']}', {sql_str(w['kr'])}, {sql_str(w['typ'])}, "
            f"{sql_num(w['gh'])}, {sql_num(w['nad'])}, {sql_dollar(w['uwagi'])}, {sql_dollar(w['wyk_zad'])});"
        )

    out_path = args.out
    if out_path is None:
        stem = args.xlsx.stem[:80]
        out_path = args.xlsx.parent / f"{stem}_czas_pracy_import.sql"

    nr_do_usuniecia = args.pracownik_nr
    if not nr_do_usuniecia and inserts:
        wb2 = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
        monthly = sorted([n for n in wb2.sheetnames if MM_RRRR.match(n.strip())])
        if monthly:
            nr_do_usuniecia = znajdz_numer_ewidencyjny(wb2[monthly[0]])
        wb2.close()
    if not nr_do_usuniecia:
        nr_do_usuniecia = "?"

    blok_delete = ""
    if not args.skip_delete and inserts and nr_do_usuniecia != "?":
        def _sk_arkusz(n: str) -> tuple[int, int]:
            a, b = n.split("-", 1)
            return (int(b), int(a))

        delete_lines: list[str] = [
            "-- Usunięcie poprzednich wpisów z tego importu: ten sam pracownik, ten sam arkusz (MM-RRRR) w znaczniku, kalendarzowy miesiąc nazwy arkusza."
        ]
        for sn in sorted(arkusze_uzyte, key=_sk_arkusz):
            zr = zakres_miesiaca_z_nazwy_arkusza(sn)
            if not zr:
                continue
            od, do = zr
            tag = f"Import: RozliczenieGodzinowe / {sn}"
            tag_q = str(tag).replace("'", "''")
            delete_lines.append(
                f"DELETE FROM public.czas_pracy_wpis\n"
                f"WHERE pracownik_nr = '{nr_do_usuniecia}'\n"
                f"  AND data >= '{od}'\n"
                f"  AND data <= '{do}'\n"
                f"  AND strpos(coalesce(uwagi, ''), '{tag_q}') > 0;\n"
            )
        blok_delete = "\n".join(delete_lines) + "\n"

    header = f"""-- Wygenerowano skryptem import_rozliczenie_godzinowe_xlsx_to_sql.py
-- Źródło: {args.xlsx.name}
-- Arkusze: miesiące MM-RRRR (bez RozliczenieKosztów / Świąt).
-- Pominięto wiersze z „WPISZ CZAS!” i puste typy/godziny.
-- Liczbowa kolumna „CZAS PRACY” → typ „praca” + godziny (KR z OBIEKT NR).
-- Kody nb-* → typy z aplikacji (urlop, zwolnienie_lekarskie, …); nb-UW bez szczegółów = 8 h.
-- Kilka wierszy „praca” w tym samym dniu, bez zakresu Od–Do, ta sama liczba godzin, różne KR → podział równy (2→½, 3→⅓).
-- Kilka KR w jednej komórce OBIEKT NR (np. 000/1079) → rozbicie na osobne wiersze i ten sam podział.
--
-- Na początku transakcji: DELETE wcześniejszych wpisów z importu (żeby ponowny import nie dublował).
-- Wyłączenie: python ... --skip-delete
--
-- Wymagane kolumny: czas-pracy-wykonywane-zadanie.sql (wykonywane_zadanie).
BEGIN;

{blok_delete}"""

    footer = "\nCOMMIT;\n"

    out_path.write_text(header + "\n".join(inserts) + footer, encoding="utf-8")
    print(f"Zapisano {len(inserts)} INSERT-ow -> {out_path}")


if __name__ == "__main__":
    main()
