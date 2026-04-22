#!/usr/bin/env python3
"""
Odczyt arkusza „Sprzet” z inwentaryzacji IT (xlsx) i zapis pliku SQL:
  INSERT ... ON CONFLICT (zewnetrzny_id) DO UPDATE ...

Wymaga: pip install openpyxl

Przykład:
  python g4-app/scripts/sprzet_generuj_sql_z_inwentaryzacji_xlsx.py ^
    "C:\\Users\\...\\2026_03_inwentaryzacja_komputerow.xlsx" ^
    g4-app/supabase/sprzet-import-wygenerowany.sql

Kolejność na Supabase:
  1) sprzet-kolumny-zewnetrzny-id-historia.sql
  2) Ten plik SQL (import)
  3) sprzet-rls-kierownik-admin-przydzial.sql
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Brak openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def sql_escape(s: str | None) -> str:
    if s is None:
        return ""
    return s.replace("'", "''")


def parse_poprzedni(raw: object) -> tuple[list[str], str | None]:
    """Z komórki Excel → lista krótkich etykiet + surowy tekst."""
    if raw is None:
        return [], None
    src = str(raw).strip()
    if not src or src.lower() == "none":
        return [], None
    parts = re.split(r"[\n\r|;\u2022]+", src)
    parts = [p.strip() for p in parts if p.strip()]
    if len(parts) == 1 and "," in parts[0]:
        parts = [p.strip() for p in parts[0].split(",") if p.strip()]
    if not parts:
        parts = [src]
    out = []
    seen = set()
    for p in parts:
        k = p.casefold()
        if k not in seen:
            seen.add(k)
            out.append(p)
    return out, src


def map_typ(kategoria: str | None) -> str:
    k = (kategoria or "").strip().lower()
    if "laptop" in k or "notebook" in k:
        return "komputer"
    if "komputer" in k or "stacjonarny" in k or "workstation" in k or "precision" in k:
        return "komputer"
    if "drukark" in k:
        return "drukarka"
    if "ksero" in k or "mf" in k:
        return "ksero"
    return "inne"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path, help="Ścieżka do pliku .xlsx")
    ap.add_argument("out_sql", type=Path, help="Ścieżka wyjściowa .sql")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, read_only=False, data_only=True)
    if "Sprzet" not in wb.sheetnames:
        print("Brak arkusza 'Sprzet'. Dostępne:", wb.sheetnames, file=sys.stderr)
        sys.exit(1)
    ws = wb["Sprzet"]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(c).strip() if c is not None else "" for c in header_row]

    lines: list[str] = [
        "-- Wygenerowano automatycznie — sprawdź typy i przypisania przed uruchomieniem na produkcji.",
        "-- Źródło: arkusz Sprzet z inwentaryzacji IT.",
        "",
        "BEGIN;",
        "",
    ]

    insert_count = 0
    for tuple_row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(tuple_row)
        d = {
            headers[i]: vals[i] if i < len(vals) else None for i in range(len(headers))
        }
        zid = d.get("ID_sprzetu") or d.get("ID sprzetu")
        if zid is None or str(zid).strip() == "":
            continue
        zew = sql_escape(str(zid).strip())

        id_uzyt = d.get("ID_uzytkownika")
        prac_nr = None
        if id_uzyt is not None and str(id_uzyt).strip() != "":
            prac_nr = sql_escape(str(id_uzyt).strip())

        stat = (d.get("Status_przypisania") or "").strip().lower()
        # Jeśli nieprzypisany — aktualny użytkownik NULL (nie ustawiamy „historycznego”).
        if "nieprzypisan" in stat:
            prac_nr = None

        kat = d.get("Kategoria_sprzetu")
        typ = map_typ(str(kat) if kat else None)

        prod = sql_escape(str(d.get("Producent") or "").strip())
        model = sql_escape(str(d.get("Model") or "").strip())
        serial = sql_escape(str(d.get("Nr_seryjny") or "").strip())
        inv = sql_escape(str(d.get("Nr_inwentarzowy") or "").strip())

        nazwa = " ".join(x for x in [prod, model] if x).strip() or zew
        if serial:
            nazwa = f"{nazwa} (SN: {serial})"

        cpu = sql_escape(str(d.get("CPU") or "").strip())
        ram = d.get("RAM_GB")
        dysk = sql_escape(str(d.get("Dysk") or "").strip())
        osys = sql_escape(str(d.get("System") or "").strip())
        akce = sql_escape(str(d.get("Akcesoria") or "").strip())
        uwagi = sql_escape(str(d.get("Uwagi") or "").strip())

        hist_raw = d.get("Uzytkownik_historyczny")
        teksty_hist, zrodlo_hist = parse_poprzedni(hist_raw)
        hist_sql_array = ", ".join(f"'{sql_escape(t)}'" for t in teksty_hist) if teksty_hist else ""

        notatki_parts = []
        if kat:
            notatki_parts.append(f"Kategoria (import): {kat}")
        if cpu:
            notatki_parts.append(f"CPU: {cpu}")
        if ram not in (None, ""):
            notatki_parts.append(f"RAM GB: {ram}")
        if dysk:
            notatki_parts.append(f"Dysk: {dysk}")
        if osys:
            notatki_parts.append(f"System: {osys}")
        if akce:
            notatki_parts.append(f"Akcesoria: {akce}")
        if uwagi:
            notatki_parts.append(f"Uwagi: {uwagi}")
        notatki = sql_escape("\n".join(notatki_parts))

        zrodlo_sql = "NULL"
        if zrodlo_hist:
            zrodlo_sql = f"'{sql_escape(zrodlo_hist)}'"

        pr_nr_sql = "NULL" if not prac_nr else f"'{prac_nr}'"

        insert_count += 1
        lines.append(
            f"""INSERT INTO public.sprzet (
  typ, nazwa, numer_inwentarzowy, data_przegladu, pracownik_nr, notatki,
  zewnetrzny_id, poprzedni_uzytkownicy_teksty, poprzedni_pracownicy_nr, poprzedni_uzytkownicy_zrodlo
) VALUES (
  '{sql_escape(typ)}',
  '{sql_escape(nazwa)}',
  {f"'{inv}'" if inv else "NULL"},
  NULL,
  {pr_nr_sql},
  '{notatki}',
  '{zew}',
  ARRAY[{hist_sql_array}]::text[],
  '{{}}'::text[],
  {zrodlo_sql}
)
ON CONFLICT ON CONSTRAINT sprzet_zewnetrzny_id_key DO UPDATE SET
  typ = EXCLUDED.typ,
  nazwa = EXCLUDED.nazwa,
  numer_inwentarzowy = EXCLUDED.numer_inwentarzowy,
  pracownik_nr = EXCLUDED.pracownik_nr,
  notatki = EXCLUDED.notatki,
  poprzedni_uzytkownicy_teksty = EXCLUDED.poprzedni_uzytkownicy_teksty,
  poprzedni_uzytkownicy_zrodlo = EXCLUDED.poprzedni_uzytkownicy_zrodlo;
"""
        )

    lines.extend(["COMMIT;", ""])
    args.out_sql.write_text("\n".join(lines), encoding="utf-8")
    print(f"Zapisano: {args.out_sql} ({insert_count} rekordów INSERT)")


if __name__ == "__main__":
    main()
